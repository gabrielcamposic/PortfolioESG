#!/usr/bin/env python3
"""
engines/B13_Cash_Parser.py — Ágora Brokerage Statement Parser

Parses account statement PDFs and XLSX spreadsheets (Extratos) from the
Ágora broker into a structured CSV of cash movements and updates the
manual_ledger.csv used by D_Publish for fund/cash balance computation.

Input:  PDF files matching *Extrato*.pdf  in Notas_Negociação/
        XLSX files matching *Extrato*.xlsx in Notas_Negociação/
Output: data/cash_movements.csv
        data/manual_ledger.csv  (consumed by D_Publish.py)

Columns (cash_movements): date, type, amount, description, source_file
Columns (manual_ledger):   date, operation, value, description

Types:
  DEPOSIT       — TED/DOC from bank into brokerage account
  WITHDRAWAL    — SPB/TED from brokerage back to bank
  DIVIDEND      — Dividend payments credited to account
  FUND_TRANSFER — Applications to external funds (e.g. Bradesco ESG)
  BOLSA         — Stock exchange settlement (buy/sell)

Usage:
    python3 engines/B13_Cash_Parser.py
"""

B13_VERSION = "2.0.0"

import csv
import json
import logging
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from engines.B11_Transactions_Parser import extract_text_from_pdf

# ── Paths ────────────────────────────────────────────────────────────────────

NOTAS_DIR = ROOT / "Notas_Negociação"
CASH_MOVEMENTS_CSV = ROOT / "data" / "cash_movements.csv"
MANUAL_LEDGER_CSV = ROOT / "data" / "manual_ledger.csv"
MANIFEST_PATH = ROOT / "data" / "processed_extratos.json"

# ── Logging ──────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - B13_Cash - %(levelname)s - %(message)s",
)
logger = logging.getLogger("B13_Cash")

# ── Constants ────────────────────────────────────────────────────────────────

MONTH_ABBR = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}

CSV_COLUMNS = ["date", "type", "amount", "description", "source_file"]


# ═════════════════════════════════════════════════════════════════════════════
# PARSING UTILITIES
# ═════════════════════════════════════════════════════════════════════════════


def parse_br_amount(s: str) -> float:
    """Parse a Brazilian-formatted number: '1.000,00' → 1000.0, '-987,88' → -987.88."""
    s = s.strip()
    sign = -1 if s.startswith("-") else 1
    s = s.lstrip("-").strip()
    s = s.replace(".", "").replace(",", ".")
    try:
        return sign * float(s)
    except ValueError:
        return 0.0


def _date_range_from_filename(filename: str) -> Tuple[Optional[int], Optional[int], Optional[int], Optional[int]]:
    """Extract (year_start, month_start, year_end, month_end) from filename.

    Patterns:
      '202501-12 Extrato Ágora.pdf'       → (2025, 1, 2025, 12)
      '202601 Extrato Ágora.pdf'           → (2026, 1, 2026, 1)
      '20260201-20260315 Extrato Ágora.pdf' → (2026, 2, 2026, 3)
    """
    base = Path(filename).stem

    # YYYYMMDD-YYYYMMDD
    m = re.match(r"(\d{4})(\d{2})(\d{2})-(\d{4})(\d{2})(\d{2})", base)
    if m:
        return int(m[1]), int(m[2]), int(m[4]), int(m[5])

    # YYYYMM-MM
    m = re.match(r"(\d{4})(\d{2})-(\d{2})", base)
    if m:
        y = int(m[1])
        return y, int(m[2]), y, int(m[3])

    # YYYYMM
    m = re.match(r"(\d{4})(\d{2})", base)
    if m:
        y, mo = int(m[1]), int(m[2])
        return y, mo, y, mo

    return None, None, None, None


def _resolve_year(month: int, y_start: Optional[int], m_start: Optional[int],
                  y_end: Optional[int], m_end: Optional[int]) -> int:
    """Determine the calendar year for a given month within a statement's range."""
    if y_start is None:
        return 2025  # safe fallback
    if y_start == y_end:
        return y_start
    # Cross-year boundary: months >= m_start belong to y_start, others to y_end
    if m_start and month >= m_start:
        return y_start
    return y_end or y_start


# ═════════════════════════════════════════════════════════════════════════════
# STATEMENT PARSER
# ═════════════════════════════════════════════════════════════════════════════


def parse_statement(text: str, source_file: str) -> List[Dict[str, Any]]:
    """Parse an Ágora brokerage account statement into cash movement entries.

    Strategy:
      1. Split text into day-level blocks using 'Saldo do dia' as delimiter.
      2. For each block, determine the date from DD/MM/YYYY or DD+MonthAbbr.
      3. Classify entries by keyword and extract amounts.

    Returns a list of dicts with keys: date, type, amount, description.
    """
    y_start, m_start, y_end, m_end = _date_range_from_filename(source_file)

    entries: List[Dict[str, Any]] = []

    # Split by "Saldo do dia" — each chunk is one day's activity
    blocks = re.split(r"Saldo do dia\s+[\d.,]+", text)

    # Track current date across blocks (some blocks lack explicit dates)
    current_date: Optional[str] = None

    for block in blocks:
        # ── Extract date ──
        date = _extract_date_from_block(block, y_start, m_start, y_end, m_end)
        if date:
            current_date = date
        if not current_date:
            continue

        # ── Deposits: TED/DOC from bank ──
        # Skip if this is a withdrawal block (#SPB#RET# = outgoing transfer)
        if ("TED BCO" in block or "DOC BCO" in block) and "#SPB#RET#" not in block:
            amt = _extract_deposit_amount(block)
            if amt and amt > 0:
                entries.append({
                    "date": current_date,
                    "type": "DEPOSIT",
                    "amount": round(amt, 2),
                    "description": "TED/DOC from bank",
                })

        # ── Withdrawals: SPB transfers back to bank ──
        # These appear in blocks with #SPB#RET# and 46663- with negative amounts
        if "#SPB#RET#" in block:
            for amt in _extract_withdrawal_amounts(block):
                entries.append({
                    "date": current_date,
                    "type": "WITHDRAWAL",
                    "amount": round(amt, 2),
                    "description": "SPB/TED to bank",
                })

        # ── Dividends ──
        if "DIVIDENDOS" in block:
            amt, ticker = _extract_dividend(block)
            if amt and amt > 0:
                entries.append({
                    "date": current_date,
                    "type": "DIVIDEND",
                    "amount": round(amt, 2),
                    "description": f"Dividendo {ticker}" if ticker else "Dividendo",
                })

        # ── Fund transfers (Bradesco ESG, etc.) ──
        for amt, fund_name in _extract_fund_transfers(block):
            entries.append({
                "date": current_date,
                "type": "FUND_TRANSFER",
                "amount": round(amt, 2),
                "description": fund_name,
            })

    return entries


def _extract_date_from_block(block: str, y_start, m_start, y_end, m_end) -> Optional[str]:
    """Find the most relevant date in a text block.

    Prefers DD/MM/YYYY explicit dates (takes the LAST one, which is typically
    the trade date rather than the TED send date).
    Falls back to a two-pass DD + month abbreviation approach that avoids
    false matches from reference numbers or amounts.
    """
    # Try explicit DD/MM/YYYY (most reliable) — take LAST occurrence
    explicit = re.findall(r"(\d{2})/(\d{2})/(\d{4})", block)
    if explicit:
        d, m, y = explicit[-1]  # last = trade date
        return f"{y}-{m}-{d}"

    # Two-pass: find month abbreviations, then find nearest valid day number
    months_found = list(re.finditer(
        r"\b(Jan|Fev|Mar|Abr|Mai|Jun|Jul|Ago|Set|Out|Nov|Dez)\b",
        block, re.IGNORECASE,
    ))
    if not months_found:
        return None

    month_match = months_found[0]
    month_num = MONTH_ABBR.get(month_match.group(1).lower())
    if not month_num:
        return None

    # Search for standalone day numbers (1-2 digits, not part of larger number)
    # in the text BEFORE the month abbreviation
    preceding = block[: month_match.start()]
    day_matches = [
        int(m.group(1))
        for m in re.finditer(r"(?<!\d)(\d{1,2})(?!\d)", preceding)
        if 1 <= int(m.group(1)) <= 31
    ]
    if day_matches:
        day = day_matches[-1]  # closest to the month abbreviation
        year = _resolve_year(month_num, y_start, m_start, y_end, m_end)
        return f"{year:04d}-{month_num:02d}-{day:02d}"

    return None


def _extract_deposit_amount(block: str) -> Optional[float]:
    """Extract deposit amount from a block containing TED/DOC keywords.

    Ágora patterns (pdfplumber text varies):
      '4666 3 - 1.000,00 REF'   (Oct 2025, TED — garbled)
      '4666- 1.000,00 REF'      (Nov 2025, DOC; Jan 2026, DOC)
      '4666- 2.000,00'           (Feb 2026, DOC)
    """
    flat = re.sub(r"\s+", " ", block)
    # Flexible: "4666" + up to 15 chars of noise + positive BR amount
    m = re.search(r"4666.{0,15}?([\d][\d.]*,\d{2})", flat)
    if m:
        amt = parse_br_amount(m.group(1))
        if amt > 0:
            return amt
    return None


def _extract_withdrawal_amounts(block: str) -> List[float]:
    """Extract withdrawal amounts from a block with #SPB#RET# pattern.

    Ágora patterns:
      '46663- -440,46'
      '46663- -1.586,69'
    """
    amounts = []
    for m in re.finditer(r"46663?-\s+-([\d.]+,\d{2})", block):
        amt = parse_br_amount(m.group(1))
        if amt > 0:
            amounts.append(-amt)  # Store as negative
    return amounts


def _extract_dividend(block: str) -> Tuple[Optional[float], Optional[str]]:
    """Extract dividend amount and ticker from a DIVIDENDOS block.

    Ágora patterns:
      'PAGAMENTO DE DIVIDENDOS S/ 10 1,25 DE VULC3'
      'PAGAMENTO DE DIVIDENDOS S/ 23 50,60 DE VULC3'

    pdfplumber may mangle the text, placing other keywords between DE and ticker.
    """
    flat = re.sub(r"\s+", " ", block)

    # Extract amount: 'DIVIDENDOS S/ [qty] [amount]'
    m = re.search(r"DIVIDENDOS\s+S/\s+\d+\s+([\d.]+,\d{2})", flat)
    if not m:
        return None, None
    amount = parse_br_amount(m.group(1))

    # Try to find ticker: standard 'DE [TICKER]' pattern
    ticker_match = re.search(r"\bDE\s+([A-Z]{3,5}\d{1,2})\b", flat)
    if not ticker_match:
        # Fallback: any BR stock symbol anywhere in the block
        ticker_match = re.search(r"\b([A-Z]{3,5}\d{1,2})\b", flat)
    ticker = ticker_match.group(1) if ticker_match else None

    return amount, ticker


def _extract_fund_transfers(block: str) -> List[Tuple[float, str]]:
    """Extract fund transfer amounts from PAG - APLIC. blocks.

    Ágora pattern:
      'PAG - APLIC. BRADESCO ESG -100,00'
      'PAG - APLIC. BRADESCO ESG -300,00'
    """
    flat = re.sub(r"\s+", " ", block)
    results = []

    for m in re.finditer(r"PAG\s*-\s*APLIC\.\s*([\w\s]+?)\s+-([\d.]+,\d{2})", flat):
        fund_name = m.group(1).strip()
        amt = -parse_br_amount(m.group(2))  # Store as negative
        results.append((amt, fund_name))

    return results


# ═════════════════════════════════════════════════════════════════════════════
# XLSX PARSER
# ═════════════════════════════════════════════════════════════════════════════


def parse_xlsx_statement(xlsx_path: Path) -> List[Dict[str, Any]]:
    """Parse an Ágora brokerage XLSX extract into manual_ledger-format entries.

    XLSX layout (sheet "Extrato Financeiro"):
      Row 7:  Headers — 'Data de referência', 'Data da liquidação', 'Lançamento',
              'Débito (R$)', [merged], 'Crédito (R$)', [merged], 'Saldo (R$)', [merged]
      Row 8+: Data rows — col A=ref_date, B=settlement_date, C=description,
              D=debit (negative), F=credit (positive), H=balance

    Returns a list of dicts: {date, operation, value, description}
    Operations: APORTE, SAQUE, BOLSA, DIVIDENDO, APLICACAO_FUNDO, RESGATE_FUNDO
    """
    import openpyxl

    entries: List[Dict[str, Any]] = []

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        logger.warning(f"  Failed to open XLSX: {xlsx_path.name} — {e}")
        return entries

    ws = wb.active or wb[wb.sheetnames[0]]

    for row_idx in range(8, ws.max_row + 1):  # Data starts at row 8 (1-indexed)
        ref_date_raw = ws.cell(row=row_idx, column=1).value
        desc_raw = ws.cell(row=row_idx, column=3).value
        debit_raw = ws.cell(row=row_idx, column=4).value
        credit_raw = ws.cell(row=row_idx, column=6).value

        # Skip header/footer/empty rows
        if not desc_raw:
            continue
        desc = str(desc_raw).strip()
        if desc in ('SALDO ANTERIOR', 'SALDO FINAL', 'Lançamentos Futuros', ''):
            continue
        if desc.startswith('Os dados acima'):
            continue

        # Parse date
        if not ref_date_raw:
            continue
        if isinstance(ref_date_raw, datetime):
            date_str = ref_date_raw.strftime("%Y-%m-%d")
        else:
            date_str = str(ref_date_raw).strip()
            if not date_str:
                continue
            # DD/MM/YYYY → YYYY-MM-DD
            m = re.match(r"(\d{2})/(\d{2})/(\d{4})", date_str)
            if m:
                date_str = f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
            else:
                continue

        # Parse amounts
        debit = float(debit_raw) if debit_raw and debit_raw != 0 else 0.0
        credit = float(credit_raw) if credit_raw and credit_raw != 0 else 0.0

        # Classify the entry
        operation, value = _classify_xlsx_entry(desc, debit, credit)
        if operation is None:
            logger.debug(f"  XLSX: Unclassified entry: {date_str} | {desc} | D={debit} C={credit}")
            continue

        entries.append({
            "date": date_str,
            "operation": operation,
            "value": round(value, 2),
            "description": desc,
        })

    wb.close()
    return entries


def _classify_xlsx_entry(desc: str, debit: float, credit: float) -> Tuple[Optional[str], float]:
    """Classify an XLSX extract line into a manual_ledger operation.

    Returns (operation, value) or (None, 0) if unclassified.
    """
    desc_upper = desc.upper()

    # Fund application: "PAG - APLIC. BRADESCO ESG ..."
    if "PAG" in desc_upper and "APLIC" in desc_upper:
        return "APLICACAO_FUNDO", abs(debit) if debit else abs(credit)

    # Fund redemption: "RESGATE" or "CREDITO FUNDO"
    if "RESGATE" in desc_upper and ("FUNDO" in desc_upper or "BRADESCO" in desc_upper):
        return "RESGATE_FUNDO", abs(credit) if credit else abs(debit)

    # Stock exchange settlement: "OPERACOES NA BOLSA"
    if "OPERACOES NA BOLSA" in desc_upper:
        # debit = bought (cash out, negative), credit = sold (cash in, positive)
        value = debit if debit else credit  # debit is already negative
        return "BOLSA", value

    # Dividend: "DIVIDENDOS" or "PAGAMENTO DE DIVIDENDOS"
    if "DIVIDENDO" in desc_upper:
        return "DIVIDENDO", abs(credit) if credit else abs(debit)

    # Deposit: TED/DOC into the account (credit side)
    if credit and credit > 0:
        if any(kw in desc_upper for kw in ("TED BCO", "DOC BCO", "TRANSF", "CTA")):
            return "APORTE", credit

    # Withdrawal: SPB/TED out of the account (debit side)
    if debit and debit < 0:
        if "#SPB#RET#" in desc_upper or "RETIRADA" in desc_upper:
            return "SAQUE", abs(debit)

    return None, 0.0


# ═════════════════════════════════════════════════════════════════════════════
# CSV OUTPUT
# ═════════════════════════════════════════════════════════════════════════════


def _load_existing_entries() -> List[Dict[str, Any]]:
    """Load existing cash_movements.csv entries."""
    if not CASH_MOVEMENTS_CSV.exists():
        return []
    entries = []
    with open(CASH_MOVEMENTS_CSV, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            entries.append(row)
    return entries


def _write_csv(entries: List[Dict[str, Any]]) -> None:
    """Write entries to cash_movements.csv, sorted by date."""
    entries.sort(key=lambda e: (e["date"], e["type"]))
    CASH_MOVEMENTS_CSV.parent.mkdir(parents=True, exist_ok=True)
    with open(CASH_MOVEMENTS_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for entry in entries:
            writer.writerow({
                "date": entry["date"],
                "type": entry["type"],
                "amount": entry["amount"],
                "description": entry.get("description", ""),
                "source_file": entry.get("source_file", ""),
            })
    logger.info(f"  Wrote {CASH_MOVEMENTS_CSV.name} ({len(entries)} entries)")


# ═════════════════════════════════════════════════════════════════════════════
# MANUAL LEDGER OUTPUT (consumed by D_Publish.py)
# ═════════════════════════════════════════════════════════════════════════════

MANUAL_LEDGER_COLUMNS = ["date", "operation", "value", "description"]


def _load_existing_manual_ledger() -> List[Dict[str, Any]]:
    """Load existing manual_ledger.csv entries."""
    if not MANUAL_LEDGER_CSV.exists():
        return []
    entries = []
    with open(MANUAL_LEDGER_CSV, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            entries.append(row)
    return entries


def _write_manual_ledger(entries: List[Dict[str, Any]]) -> None:
    """Write entries to manual_ledger.csv, sorted by date."""
    entries.sort(key=lambda e: (e["date"], e.get("operation", "")))
    MANUAL_LEDGER_CSV.parent.mkdir(parents=True, exist_ok=True)
    with open(MANUAL_LEDGER_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=MANUAL_LEDGER_COLUMNS)
        writer.writeheader()
        for entry in entries:
            writer.writerow({
                "date": entry["date"],
                "operation": entry["operation"],
                "value": entry["value"],
                "description": entry.get("description", ""),
            })
    logger.info(f"  Wrote {MANUAL_LEDGER_CSV.name} ({len(entries)} entries)")


# ═════════════════════════════════════════════════════════════════════════════
# MANIFEST (idempotent processing)
# ═════════════════════════════════════════════════════════════════════════════


def _load_manifest() -> dict:
    if MANIFEST_PATH.exists():
        try:
            with open(MANIFEST_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError):
            pass
    return {"processed": [], "last_run": None}


def _save_manifest(manifest: dict) -> None:
    MANIFEST_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(MANIFEST_PATH, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)


# ═════════════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════════════


def main() -> int:
    logger.info(f"B13_Cash_Parser v{B13_VERSION} — Parsing brokerage statements")

    if not NOTAS_DIR.exists():
        logger.warning(f"  Notas directory not found: {NOTAS_DIR}")
        return 0

    manifest = _load_manifest()
    processed_set = set(manifest.get("processed", []))

    all_csv_entries: List[Dict[str, Any]] = []  # For cash_movements.csv
    all_ledger_entries: List[Dict[str, Any]] = []  # For manual_ledger.csv
    new_count = 0

    # ── Phase 1: Process PDF Extracts ────────────────────────────────────────
    pdfs = sorted(NOTAS_DIR.glob("*Extrato*.pdf"))
    logger.info(f"  Found {len(pdfs)} Extrato PDF(s)")

    for pdf in pdfs:
        key = pdf.name
        if key in processed_set:
            logger.info(f"  Skipping (already processed): {key}")
            continue

        logger.info(f"  Processing PDF: {key}")
        text = extract_text_from_pdf(pdf)
        if not text or not text.strip():
            logger.warning(f"  Failed to extract text from: {key}")
            continue

        entries = parse_statement(text, key)
        for e in entries:
            e["source_file"] = key

        logger.info(f"    Found {len(entries)} cash movement(s)")
        for e in entries:
            logger.info(f"    {e['date']}  {e['type']:15s}  {e['amount']:>10.2f}  {e['description']}")

        all_csv_entries.extend(entries)
        processed_set.add(key)
        new_count += 1

    # ── Phase 2: Process XLSX Extracts ───────────────────────────────────────
    xlsx_files = sorted(NOTAS_DIR.glob("*Extrato*.xlsx"))
    logger.info(f"  Found {len(xlsx_files)} Extrato XLSX file(s)")

    # XLSX files are always re-parsed (they're the source of truth for
    # manual_ledger.csv). We use the manifest only to track which XLSX
    # files contributed, but always re-read to catch updated exports.
    xlsx_ledger_entries: List[Dict[str, Any]] = []

    for xlsx_path in xlsx_files:
        key = xlsx_path.name
        logger.info(f"  Processing XLSX: {key}")

        entries = parse_xlsx_statement(xlsx_path)
        logger.info(f"    Found {len(entries)} ledger entries")
        for e in entries:
            logger.info(f"    {e['date']}  {e['operation']:20s}  {e['value']:>10.2f}  {e['description'][:60]}")

        xlsx_ledger_entries.extend(entries)

        if key not in processed_set:
            processed_set.add(key)
            new_count += 1

    # ── Phase 3: Write manual_ledger.csv from XLSX data ──────────────────────
    if xlsx_ledger_entries:
        # Deduplicate: same (date, operation, value) = same entry
        seen_ledger = set()
        unique_ledger = []
        for e in xlsx_ledger_entries:
            k = (e["date"], e["operation"], str(e["value"]))
            if k not in seen_ledger:
                seen_ledger.add(k)
                unique_ledger.append(e)
            else:
                logger.info(f"    Dedup ledger: skipped {k}")

        _write_manual_ledger(unique_ledger)
        all_ledger_entries = unique_ledger
    else:
        logger.info("  No XLSX extracts found — manual_ledger.csv unchanged")

    # ── Phase 4: Write cash_movements.csv from PDF data ──────────────────────
    if not all_csv_entries and not pdfs:
        logger.info("  No Extrato PDFs found")
    elif not all_csv_entries:
        logger.info("  No new PDFs to process")
        if not CASH_MOVEMENTS_CSV.exists():
            logger.info("  Reprocessing all PDFs to generate CSV...")
            for pdf in pdfs:
                if pdf.name in processed_set or True:  # force reprocess
                    text = extract_text_from_pdf(pdf)
                    if text and text.strip():
                        entries = parse_statement(text, pdf.name)
                        for e in entries:
                            e["source_file"] = pdf.name
                        all_csv_entries.extend(entries)

    if all_csv_entries:
        # Deduplicate
        seen = set()
        unique = []
        for e in all_csv_entries:
            k = (e["date"], e["type"], e["amount"])
            if k not in seen:
                seen.add(k)
                unique.append(e)
        all_csv_entries = unique

        # Merge with existing
        existing = _load_existing_entries()
        for e in existing:
            k = (e["date"], e["type"], e["amount"])
            if k not in seen:
                all_csv_entries.append(e)
                seen.add(k)

        _write_csv(all_csv_entries)

    # Save manifest
    manifest["processed"] = sorted(list(processed_set))
    manifest["last_run"] = datetime.now().isoformat()
    _save_manifest(manifest)
    logger.info(f"  Done. Processed {new_count} new file(s), "
                f"{len(all_csv_entries)} CSV entries, "
                f"{len(all_ledger_entries)} ledger entries.")

    return 0


if __name__ == "__main__":
    sys.exit(main())

